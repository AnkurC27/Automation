from bs4 import BeautifulSoup
import requests
import datetime
import pandas as pd
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from urllib.parse import urlparse 
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoAlertPresentException
import os
from PIL import Image, ImageDraw, ImageFont
import time

# Set the path to the EdgeDriver executable
edgedriver_path = r'C:\Users\ankur.chadha\desktop\msedgedriver'

# Set up Edge options
edge_options = webdriver.EdgeOptions()

# Add the Edge driver directory to the PATH environment variable
os.environ["PATH"] += os.pathsep + edgedriver_path

# Initialize the Edge driver
driver = webdriver.Edge(options=edge_options)

# Maximize the browser window
driver.maximize_window()

# Read manufacturer codes from Excel file
excel_file_path = r'C:\Users\ankur.chadha\Desktop\Automation\skutest.xlsx'
sku_test_df = pd.read_excel(excel_file_path)

# List of websites
websites = ['https://www.homedepot.com/s/{model_number}?NCNI-5',
            'https://www.motion.com/products/search;q={model_number};origin=search',
            'https://usatoolsinc.com/search.php?search_query=%E2%80%8E{model_number}&section=product',
            'https://www.grainger.com/search?searchQuery={model_number}&searchBar=true',
            'https://www.coastaltool.com/search?type=article%2Cpage%2Cproduct&q={model_number}*']

# Website Names
website_names = {
    'www.homedepot.com': 'home_depot',
    'www.motion.com': 'motion_i',
    'www.grainger.com': 'grainger',
    'www.usatoolsinc.com': 'usa_tools',
    'www.coastaltool.com': 'coastal'
}

# Get the column index of the 'Model Number' header
model_number_col_index = sku_test_df.columns.get_loc('Model Number 1')
model_number_col_index_2 = sku_test_df.columns.get_loc('Model Number 2')

# Get the column index of the 'Item Number' header
item_number_col_index = sku_test_df.columns.get_loc('Item Number')
item_desc_col_index = sku_test_df.columns.get_loc('Item Description')

website_columns = {
    'homedepot.com': 'Home_Depot_Link',
    'motion.com': 'Motion_i_Link',
    'grainger.com': 'Grainger_Link',
    'usatoolsinc.com': 'USA_Tools_Link',
    'coastaltool.com': 'Coastal_Link'
}

price_columns = {
    'homedepot.com': 'Home_Depot_Price',
    'motion.com': 'Motion_Price',
    'grainger.com': 'Grainger_Price',
    'usatoolsinc.com': 'USA_Tools_Price',
    'coastaltool.com': 'Coastal_Price'
}
#find the price elements on each website
def get_price_from_html(html_content, website_name):

    try:
        soup = BeautifulSoup(html_content, 'html.parser')
    except Exception as e:
        print(f"Error parsing HTML: {str(e)}")
        return None

    #Find the price based on website
    try: 
        if website_name == 'www.homedepot.com':
            price_tag = soup.find('div', class_='price')
        elif website_name == 'www.motion.com':
            price_tag = soup.find('div', class_='price me-4p price-style ng-star-inserted')
        elif website_name == 'usatoolsinc.com':
            price_tag = soup.find('span', class_='price price--withoutTax')
        elif website_name == 'www.grainger.com':
            price_tag = soup.find('span', class_='rbqUOE lVwVq5')
        elif website_name == 'www.coastaltool.com':
            price_tag = soup.find('div', class_='price--main')
        else:
            price_tag = None
    except Exception as e:
        print(f"Error finding price tag in HTML: {str(e)}")
        return None

    if price_tag is not None:
        try:
            return price_tag.text
        except Exception as e:
            print(f"Error getting text from price tag: {str(e)}")
            return None
    else:
        return None                        

#watermark that adds a white border and a yellow highlight 
def add_watermark(screenshot_filename, item_number, item_desc):
    img = Image.open(screenshot_filename)
    draw = ImageDraw.Draw(img)
    date_time = datetime.datetime.now().strftime("%m/%d/%Y, %H:%M:%S")
    watermark_text = f"{item_number}, {item_desc}, {date_time}"
    position = (10, 10)
    font = ImageFont.truetype("C:\\Windows\\Fonts\\arial.ttf", 30)
    color = "black" 
    stroke_color = "white"
    stroke_width = 2
    background_color = "yellow"

    # Get image and text dimensions
    img_width, img_height = img.size
    text_width, text_height = draw.textsize(watermark_text, font=font)

    position = (img_width - text_width - 10, img_height - text_height -10)
    padding = 10

    draw.rectangle([position[0]-padding, position[1]-padding, position[0]+text_width+padding, position[1]+text_height+padding], fill=background_color)

    draw.text(position, watermark_text, font=font, fill=color, stroke_width=stroke_width, stroke_fill=stroke_color)
    img.save(screenshot_filename)


# Iterate over the rows in the Excel file
for index, row in sku_test_df.iterrows():
    model_number_1 = row[model_number_col_index]
    model_number_2 = row[model_number_col_index_2]

    item_number = row[item_number_col_index]
    item_desc = str(row[item_desc_col_index])
    if pd.isnull(item_number):
        break

    folder_name = str(int(item_number/100)*100)

    if not os.path.exists(folder_name):
        try:
            os.mkdir(folder_name)
            print(f"Directory '{folder_name}' created.")
        except Exception as e:
            print(f"Could not create the directory. Error: {str(e)}")

    for model_number in [model_number_1, model_number_2]:
        # Check if model number is not null before opening websites and taking screenshots
        if not model_number or model_number.isspace():
            continue

        website_wait_times = {
            'https://www.homedepot.com/s/{model_number}?NCNI-5': 6,
            'https://www.motion.com/products/search;q={model_number};origin=search': 3,
            'https://www.grainger.com/search?searchQuery={model_number}&searchBar=true': 4,
            'https://usatoolsinc.com/search.php?search_query=%E2%80%8E{model_number}&section=product': 3,
            'https://www.coastaltool.com/search?type=article%2Cpage%2Cproduct&q={model_number}*': 3
        }

        for website in websites:
            wait_time = website_wait_times.get(website, 2)
            time.sleep(wait_time)

            website_url = website.format(model_number=model_number)
            driver.get(website_url)

            try:
                driver.switch_to.alert.accept()
            except NoAlertPresentException:
                pass

            parsed_uri = urlparse(website_url)
            website_name = '{uri.netloc}'.format(uri=parsed_uri)
            website_name = website_name.replace("www.", "")  # remove www. part

            try:
                html_content = driver.page_source
                price = get_price_from_html(html_content, website_name)
            except Exception as e:
                print(f"Error getting price from website {website}: {str(e)}")
                price = None

            item_desc = str(row[item_desc_col_index])
            custom_website_name = website_names.get(website_name, website_name)
            item_desc = str(row[item_desc_col_index]).replace('\'', '_').replace('\"', '_')

            screenshot_column_name = website_columns[website_name]
            sku_test_df.at[index, screenshot_column_name] = website_url

            price_column_name = price_columns[website_name]
            sku_test_df.at[index, price_column_name] = price

            # Take a screenshot
            screenshot_filename = f'{folder_name}/{row["Item Number"]}_{item_desc}_{custom_website_name}_{index}.png'
            driver.save_screenshot(screenshot_filename)
            add_watermark(screenshot_filename, item_number, item_desc)
            driver.delete_all_cookies()


from openpyxl import load_workbook

# Load existing workbook
book = load_workbook(excel_file_path)

# Select the sheet
writer = book['Sheet1']

# Write to the excel file
for column in sku_test_df.columns:
    col_index = sku_test_df.columns.get_loc(column)
    for idx, value in sku_test_df[column].items():
        writer.cell(row=idx+2, column=col_index+1, value=value)  # indices are 1-based in openpyxl

# Save the workbook
book.save(excel_file_path)

# Close the workbook
book.close()

# Close the web driver
driver.quit()